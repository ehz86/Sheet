from openpyxl import workbook
import openpyxl

wb = openpyxl.load_workbook('данные для задачки.xlsx')
from exception1 import Exp1
from exception2 import Exp2
from exception3 import Exp3
from exception4 import Exp4
from exception5 import Exp5
from exception6 import Exp6
ws2 = wb['info']
ws1 = wb['Лист3']
n = ws1.max_row+1
m = ws2.max_row+1
i = 1
for h2 in range(1,m):
    for h1 in range(1,n):
        a1 = ws1.cell(row=h1, column=1).value
        b1 = Exp1(a1)
        a2 = ws1.cell(row=h1, column=2).value
        b2 = Exp2(a2)
        a3 = ws1.cell(row=h1, column=3).value
        b3 = Exp3(a3)
        a4 = ws1.cell(row=h1, column=4).value
        b4 = Exp4(a4)
        a5 = ws1.cell(row=h1, column=5).value
        b5 = Exp5(a5)
        a6 = ws1.cell(row=h1, column=6).value
        b6 = Exp6(a6)
        # 1 ячейка
        if ws2.cell(row=h2, column=1).value == ws1.cell(row=h1, column=1).value or ws1.cell(row=h1, column=1).value == '#' or(b1 is not  None and ws2.cell(row=h2, column=1).value != b1):
            # 2 ячейка
            if ws2.cell(row=h2, column=2).value == ws1.cell(row=h1, column=2).value or ws1.cell(row=h1, column=2).value == '#' or(b2 is not  None and  ws2.cell(row=h2, column=2).value != b2):
                # 3 ячейка
                if ws2.cell(row=h2, column=3).value == ws1.cell(row=h1, column=3).value or ws1.cell(row=h1,column=3).value == '#' or(b3 is not None and  ws2.cell(row=h2, column=3).value != b3):
                    # 4 ячейка
                    if ws2.cell(row=h2, column=4).value == ws1.cell(row=h1, column=4).value or ws1.cell(row=h1,column=4).value == '#' or(b4 is not  None and  ws2.cell(row=h2, column=4).value != b4):
                        # 5 ячейка
                        if ws2.cell(row=h2, column=5).value == ws1.cell(row=h1, column=5).value or ws1.cell(row=h1,column=5).value == '#' or (b5 is not None and ws2.cell(row=h2, column=5).value != b5):
                            # 6 ячейка
                            if ws2.cell(row=h2, column=6).value == ws1.cell(row=h1, column=6).value or ws1.cell(row=h1,column=6).value == '#' or (b6 is not None and ws2.cell(row=h2, column=6).value != b6):
                                # Строка совпала
                                ws2.cell(row=h2, column=7).value = ws1.cell(row=h1, column=7).value
        if ws2.cell(row=h2, column=7).value == None:
            #Строка не совпала
            ws2.cell(row=h2, column=7).value = 0

wb.save('данные для задачки.xlsx')
